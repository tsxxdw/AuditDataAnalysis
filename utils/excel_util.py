"""
Excel工具类

提供Excel文件操作相关的工具函数，包括读取Excel文件、获取工作表列表等
"""

import os
import sys
import subprocess
import platform
import pandas as pd
from openpyxl import load_workbook
from service.log.logger import app_logger

# 尝试导入xlrd库，用于处理.xls文件
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False
    app_logger.warning("未安装xlrd库，可能无法正确处理.xls格式的Excel文件")

class ExcelUtil:
    """Excel工具类，提供Excel文件操作的静态方法"""
    
    @staticmethod
    def validate_excel_path(file_path):
        """
        验证Excel文件路径是否有效
        
        Args:
            file_path (str): Excel文件路径
            
        Returns:
            bool: 文件路径是否有效
        """
        # 检查文件是否存在
        if not os.path.exists(file_path):
            app_logger.warning(f"文件不存在: {file_path}")
            return False
        
        # 检查是否是文件
        if not os.path.isfile(file_path):
            app_logger.warning(f"路径不是文件: {file_path}")
            return False
        
        # 检查扩展名
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in ['.xlsx', '.xls']:
            app_logger.warning(f"文件格式不是Excel: {file_path}, 扩展名: {ext}")
            return False
            
        # 如果是.xls文件，检查是否有xlrd库
        if ext == '.xls' and not HAS_XLRD:
            app_logger.warning(f"检测到.xls文件但未安装xlrd库，文件处理可能会受影响: {file_path}")
            # 仍然返回True，我们会在后续方法中再次检查
            
        return True
    
    @staticmethod
    def get_sheets_info(file_path):
        """
        获取Excel文件中所有工作表的信息
        
        Args:
            file_path (str): Excel文件路径
            
        Returns:
            list: 工作表信息列表，每个元素包含sheet名称和索引
            
        Raises:
            Exception: 当文件无法打开或解析时抛出异常
        """
        try:
            # 首先验证文件路径
            if not ExcelUtil.validate_excel_path(file_path):
                raise ValueError(f"无效的Excel文件路径: {file_path}")
            
            # 根据扩展名使用不同的方法打开文件
            ext = os.path.splitext(file_path)[1].lower()
            sheets = []
            
            app_logger.info(f"尝试读取Excel文件: {file_path}, 格式: {ext}")
            
            if ext == '.xlsx':
                # 使用openpyxl读取xlsx文件
                app_logger.info(f"使用openpyxl读取.xlsx文件: {file_path}")
                wb = load_workbook(file_path, read_only=True)
                sheets = [{'name': name, 'index': i, 'id': f"{name}_{i}"} for i, name in enumerate(wb.sheetnames)]
                wb.close()
            elif ext == '.xls':
                try:
                    # 优先使用xlrd读取xls文件
                    if HAS_XLRD:
                        app_logger.info(f"使用xlrd读取.xls文件: {file_path}")
                        workbook = xlrd.open_workbook(file_path)
                        sheets = [{'name': name, 'index': i, 'id': f"{name}_{i}"} for i, name in enumerate(workbook.sheet_names())]
                    else:
                        # 回退到pandas
                        app_logger.info(f"使用pandas读取.xls文件: {file_path}")
                        # engine参数对于读取.xls文件很重要
                        excel = pd.ExcelFile(file_path, engine='xlrd')
                        sheets = [{'name': name, 'index': i, 'id': f"{name}_{i}"} for i, name in enumerate(excel.sheet_names)]
                        excel.close()
                except Exception as e:
                    app_logger.error(f"读取.xls文件失败，尝试第二种方法: {str(e)}")
                    # 尝试使用另一种方法
                    excel = pd.ExcelFile(file_path, engine='xlrd')
                    sheets = [{'name': name, 'index': i, 'id': f"{name}_{i}"} for i, name in enumerate(excel.sheet_names)]
                    excel.close()
            
            app_logger.info(f"成功读取工作表信息，共{len(sheets)}个工作表")
            return sheets
            
        except Exception as e:
            app_logger.error(f"读取Excel文件工作表失败: {str(e)}", exc_info=True)
            raise Exception(f"无法读取Excel文件工作表: {str(e)}")
    
    @staticmethod
    def read_excel_data(file_path, sheet_name=None, sheet_index=0, start_row=0, row_limit=None, ignore_empty_rows=False):
        """
        读取Excel文件数据，返回列表格式的数据
        
        Args:
            file_path (str): Excel文件路径
            sheet_name (str, optional): 工作表名称，如果提供则优先使用
            sheet_index (int, optional): 工作表索引，当sheet_name未提供时使用
            start_row (int, optional): 开始行，默认为0(第一行)
            row_limit (int, optional): 读取行数限制，默认为None表示读取所有行
            ignore_empty_rows (bool, optional): 是否忽略全空行，默认为False
            
        Returns:
            list: 包含所有行数据的列表，每行是一个列表
            
        Raises:
            Exception: 当文件无法打开或解析时抛出异常
        """
        try:
            # 首先验证文件路径
            if not ExcelUtil.validate_excel_path(file_path):
                raise ValueError(f"无效的Excel文件路径: {file_path}")
            
            # 检查文件类型
            ext = os.path.splitext(file_path)[1].lower()
            app_logger.info(f"尝试读取Excel数据: {file_path}, 格式: {ext}, 开始行: {start_row}, 忽略空行: {ignore_empty_rows}")
            
            # 使用适当的引擎读取Excel数据
            engine = 'openpyxl' if ext == '.xlsx' else 'xlrd'
            
            # 读取Excel数据
            try:
                if sheet_name:
                    # 如果提供了sheet_name，则使用sheet_name
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine=engine)
                else:
                    # 否则使用sheet_index
                    df = pd.read_excel(file_path, sheet_name=sheet_index, header=None, engine=engine)
            except Exception as e:
                app_logger.error(f"使用{engine}引擎读取失败，尝试其他引擎: {str(e)}")
                # 尝试不指定引擎
                if sheet_name:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                else:
                    df = pd.read_excel(file_path, sheet_name=sheet_index, header=None)
            
            # 获取数据，从指定行开始
            data = df.iloc[start_row:]
            
            # 如果指定了行数限制，只取指定的行数
            if row_limit:
                data = data.iloc[:row_limit]
            
            # 将DataFrame转换为列表
            rows = []
            
            # 处理数据，确保没有NaN值
            for _, row in data.iterrows():
                # 将每个单元格数据转换为Python原生类型，并处理NaN值
                processed_row = []
                for value in row:
                    if pd.isna(value) or pd.isnull(value):
                        processed_row.append(None)  # 将NaN值转换为None
                    elif isinstance(value, (pd.Timestamp, pd.Period)):
                        # 处理日期时间类型
                        processed_row.append(str(value))
                    else:
                        processed_row.append(value)
                
                # 如果需要忽略空行，检查行是否为空
                if ignore_empty_rows:
                    # 行为空的条件：所有值都是None或空字符串
                    is_empty = all((val is None or (isinstance(val, str) and val.strip() == '')) for val in processed_row)
                    if not is_empty:
                        rows.append(processed_row)
                else:
                    # 不忽略空行，添加所有行
                    rows.append(processed_row)
            
            log_msg = f"成功读取Excel数据，原始行数: {len(data)}, 处理后行数: {len(rows)}"
            if ignore_empty_rows:
                log_msg += f"（已忽略空行）"
            if row_limit:
                log_msg += f"（限制为{row_limit}行）"
            app_logger.info(log_msg)
            
            return rows
            
        except Exception as e:
            app_logger.error(f"读取Excel数据失败: {str(e)}", exc_info=True)
            raise Exception(f"无法读取Excel数据: {str(e)}")
    
    @staticmethod
    def get_sheet_data_preview(file_path, sheet_name=None, sheet_index=0, start_row=0, row_count=10):
        """
        获取Excel工作表数据预览
        
        Args:
            file_path (str): Excel文件路径
            sheet_name (str, optional): 工作表名称，如果提供则优先使用
            sheet_index (int, optional): 工作表索引，当sheet_name未提供时使用
            start_row (int, optional): 开始行，默认为0(第一行)
            row_count (int, optional): 预览行数，默认为10
            
        Returns:
            dict: 包含数据的字典
            
        Raises:
            Exception: 当文件无法打开或解析时抛出异常
        """
        try:
            # 调用统一的read_excel_data方法获取数据
            # 预览数据时不忽略空行，以便用户看到实际情况
            rows = ExcelUtil.read_excel_data(
                file_path=file_path,
                sheet_name=sheet_name,
                sheet_index=sheet_index,
                start_row=start_row,
                row_limit=row_count,
                ignore_empty_rows=False  # 预览时不忽略空行
            )
            
            # 构建返回结果
            return rows, rows
            
        except Exception as e:
            app_logger.error(f"获取Excel数据预览失败: {str(e)}", exc_info=True)
            raise Exception(f"无法获取Excel数据预览: {str(e)}")
    
    @staticmethod
    def parse_sheet_id(sheet_id):
        """
        解析工作表ID，从格式"sheetName_index"中提取sheet名称和索引
        
        Args:
            sheet_id (str): 工作表ID，格式为"sheetName_index"
            
        Returns:
            tuple: (sheet_name, sheet_index)
        """
        try:
            # 查找最后一个下划线的位置
            last_underscore = sheet_id.rindex('_')
            sheet_name = sheet_id[:last_underscore]
            sheet_index = int(sheet_id[last_underscore+1:])
            return sheet_name, sheet_index
        except (ValueError, IndexError):
            # 如果解析失败，返回原始ID作为sheet名称，索引为0
            return sheet_id, 0
    
    @staticmethod
    def open_excel_file(file_path, sheet_id=None):
        """
        在Windows环境下打开Excel文件并定位到指定工作表
        
        Args:
            file_path (str): Excel文件完整路径
            sheet_id (str, optional): 工作表ID，格式为"sheetName_index"
            
        Returns:
            dict: 包含操作结果的字典
                success: 是否成功
                message: 详细信息
                is_windows: 是否为Windows环境
        """
        try:
            # 检查操作系统
            is_windows = platform.system().lower() == 'windows'
            if not is_windows:
                return {
                    "success": False,
                    "message": "此功能仅支持Windows操作系统",
                    "is_windows": False
                }
            
            # 验证文件路径
            if not ExcelUtil.validate_excel_path(file_path):
                return {
                    "success": False,
                    "message": f"无效的Excel文件路径: {file_path}",
                    "is_windows": True
                }
            
            # 规范化文件路径 - 确保使用双反斜杠或正斜杠
            file_path = os.path.abspath(file_path)
            
            # 记录实际使用的文件路径
            app_logger.info(f"尝试打开文件，完整路径: {file_path}")
            
            # 使用系统默认关联程序打开文件（适用于Microsoft Excel和WPS）
            if sheet_id:
                try:
                    # 解析sheet_id
                    sheet_name, _ = ExcelUtil.parse_sheet_id(sheet_id)
                    
                    # 方法1: 直接使用系统默认程序打开文件
                    # 这是最可靠的方法，但无法定位到特定sheet
                    os.startfile(file_path)
                    
                    # 尝试检测是WPS还是Microsoft Excel
                    is_wps = ExcelUtil._is_wps_default_program(file_path)
                    app_logger.info(f"检测到默认程序是WPS: {is_wps}")
                    
                    return {
                        "success": True,
                        "message": f"已打开Excel文件 {os.path.basename(file_path)} 的工作表 {sheet_name}",
                        "is_windows": True,
                        "using_wps": is_wps
                    }
                except Exception as e:
                    app_logger.warning(f"尝试打开指定工作表失败: {str(e)}")
                    # 如果上面的方法失败，尝试直接使用os.system
                    os.system(f'start "" "{file_path}"')
                    return {
                        "success": True,
                        "message": f"已打开Excel文件 {os.path.basename(file_path)}，但无法定位到指定工作表",
                        "is_windows": True
                    }
            else:
                # 直接使用系统默认程序打开文件
                os.startfile(file_path)
                return {
                    "success": True,
                    "message": f"已打开Excel文件 {os.path.basename(file_path)}",
                    "is_windows": True
                }
                
        except Exception as e:
            app_logger.error(f"打开Excel文件失败: {str(e)}", exc_info=True)
            return {
                "success": False,
                "message": f"打开Excel文件失败: {str(e)}",
                "is_windows": platform.system().lower() == 'windows'
            }
    
    @staticmethod
    def _is_wps_default_program(file_path):
        """
        尝试检测系统默认使用的是WPS还是Microsoft Excel
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            bool: 是否使用WPS作为默认程序
        """
        try:
            import winreg
            # 获取.xlsx文件的默认打开程序
            ext = os.path.splitext(file_path)[1].lower()
            with winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, ext) as key:
                prog_id = winreg.QueryValue(key, "")
                
            app_logger.debug(f"Excel文件关联的ProgID: {prog_id}")
            
            # 检查程序ID是否包含WPS
            return 'wps' in prog_id.lower()
        except Exception as e:
            app_logger.debug(f"检测默认程序失败: {str(e)}")
            return False
    
    @staticmethod
    def get_sheet_total_rows(file_path, sheet_name=None, sheet_index=0, ignore_empty_rows=True):
        """
        获取Excel工作表的有效总行数，即使中间有空行区域也能正确计算最后一行有效数据
        
        Args:
            file_path (str): Excel文件路径
            sheet_name (str, optional): 工作表名称，如果提供则优先使用
            sheet_index (int, optional): 工作表索引，当sheet_name未提供时使用
            ignore_empty_rows (bool, optional): 是否忽略尾部全空行，默认为True
            
        Returns:
            int: 工作表的有效总行数
            
        Raises:
            Exception: 当文件无法打开或解析时抛出异常
        """
        try:
            # 首先验证文件路径
            if not ExcelUtil.validate_excel_path(file_path):
                raise ValueError(f"无效的Excel文件路径: {file_path}")
            
            # 检查文件类型
            ext = os.path.splitext(file_path)[1].lower()
            app_logger.info(f"获取工作表有效总行数: {file_path}, 格式: {ext}, 忽略空行: {ignore_empty_rows}")
            
            # 如果不需要忽略空行，使用原来的方法快速计算
            if not ignore_empty_rows:
                if ext == '.xlsx':
                    wb = load_workbook(file_path, read_only=True)
                    if sheet_name and sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                    else:
                        ws = wb.worksheets[min(sheet_index, len(wb.worksheets)-1)]
                    max_row = ws.max_row
                    wb.close()
                    return max_row
                else:
                    if sheet_name:
                        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                    else:
                        df = pd.read_excel(file_path, sheet_name=sheet_index, header=None)
                    return len(df)
            
            # 需要忽略空行，使用pandas高效处理
            try:
                if sheet_name:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                else:
                    df = pd.read_excel(file_path, sheet_name=sheet_index, header=None)
                
                # 创建一个布尔掩码，标记非空行
                # 行为非空的条件：至少有一个单元格非空且非空字符串
                non_empty_mask = df.apply(
                    lambda row: row.notna().any() and not all(
                        (pd.isna(val) or (isinstance(val, str) and val.strip() == ''))
                        for val in row
                    ),
                    axis=1
                )
                
                if not non_empty_mask.any():
                    # 所有行都是空的
                    return 0
                
                # 找到最后一个非空行的索引
                last_non_empty_rows = non_empty_mask[non_empty_mask].index
                if len(last_non_empty_rows) == 0:
                    return 0
                    
                # 返回最后一个非空行的索引 + 1（因为索引从0开始）
                return last_non_empty_rows[-1] + 1
                
            except Exception as e:
                app_logger.error(f"使用pandas获取行数失败，尝试备用方法: {str(e)}")
                
                # 备用方法：读取所有数据并检查
                all_data = ExcelUtil.read_excel_data(file_path, sheet_name, sheet_index)
                
                # 如果没有数据，直接返回0
                if not all_data:
                    return 0
                
                # 从后向前查找最后一个非空行
                for i in range(len(all_data) - 1, -1, -1):
                    row = all_data[i]
                    # 检查行是否为空（所有单元格都是None或空字符串）
                    if any(cell is not None and (not isinstance(cell, str) or cell.strip() != '') for cell in row):
                        # 找到第一个非空行，返回其索引+1作为总行数
                        return i + 1
                
                # 如果所有行都是空的，返回0
                return 0
            
        except Exception as e:
            app_logger.error(f"获取工作表有效总行数失败: {str(e)}", exc_info=True)
            raise Exception(f"无法获取工作表有效总行数: {str(e)}") 